# Studentクラスをインポートする
from student.models import Student
# JsonResponseモジュールをインポートする
from django.http import JsonResponse
# jsonモジュールをインポートする
import json
# Q検索をインポートする
from django.db.models import Q
# uuidクラスをインポートする
import uuid
# ハッシュライブラリをインポートする
import hashlib
# 設定をインポートする
from django.conf import settings
# osをインポートする
import os
# Excel処理モジュールをインポートする
import openpyxl
# ここでビューを作成する


def get_students(request):
    """すべての学生の情報を取得する"""
    try:
        # ORMを使用してすべての学生情報を取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.all().values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code':1, 'data':students})
    except Exception as e:
        # 例外が発生した場合は、戻り値
        return JsonResponse({'code': 0, 'msg': "学生情報の取得時に例外が発生しました。詳細：" + str(e)})


def query_students(request):
    """学生情報を検索する"""
    # 送信された検索条件を受け取る--- axiosはデフォルトでjson --- 辞書型（'inputstr'）-- data['inputstr']
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 条件に合致する学生情報をORMを使用して取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.filter(Q(sno__icontains=data['inputstr']) | Q(name__icontains=data['inputstr']) |
                                              Q(gender__icontains=data['inputstr']) | Q(mobile__icontains=data['inputstr'])
                                              | Q(email__icontains=data['inputstr']) | Q(address__icontains=data['inputstr'])).values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        # 例外が発生した場合は、戻り値
        return JsonResponse({'code': 0, 'msg': "学生情報の検索時に例外が発生しました。詳細：" + str(e)})


def is_exists_sno(request):
    """学籍番号が存在するかどうかを判断する"""
    # 送られてきた学籍番号を受け取る
    data = json.loads(request.body.decode('utf-8'))
    # 検証を行う
    try:
        obj_students = Student.objects.filter(sno=data['sno'])
        if obj_students.count() == 0:
            return JsonResponse({'code': 1, 'exists': False})
        else:
            return JsonResponse({'code': 1, 'exists': True})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"学籍番号の検証に失敗しました。具体的な原因：" + str(e)})


def add_student(request):
    """データベースに学生を追加する"""
    # フロントエンドから送られてきた値を受け取る
    data = json.loads(request.body.decode("utf-8"))
    try:
        # データベースに追加する
        obj_student = Student(sno=data['sno'],name=data['name'],gender=data['gender'],
                              birthday=data['birthday'],mobile=data['mobile'],
                              email= data['email'], address=data['address'],image=data['image'])
        # 追加を実行する
        obj_student.save()
        # ORMを使用してすべての学生情報を取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.all().values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "データベースへの追加時に例外が発生しました。具体的な原因：" + str(e)})


def update_student(request):
    """データベースの学生情報を更新する"""
    # フロントエンドから送られてきた値を受け取る
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 更新する学生情報を検索する
        obj_student = Student.objects.get(sno=data['sno'])
        # 順番に更新する
        obj_student.name = data['name']
        obj_student.gender = data['gender']
        obj_student.birthday = data['birthday']
        obj_student.mobile = data['mobile']
        obj_student.email = data['email']
        obj_student.address = data['address']
        obj_student.image = data['image']
        # 保存する
        obj_student.save()
        # ORMを使用してすべての学生情報を取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.all().values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "データベースへの更新保存時に例外が発生しました。具体的な原因：" + str(e)})


def delete_student(request):
    """1件の学生情報を削除する"""
    # フロントエンドから送られてきた値を受け取る
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 削除する学生情報を検索する
        obj_student = Student.objects.get(sno=data['sno'])
        # 削除する
        obj_student.delete()
        # ORMを使用してすべての学生情報を取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.all().values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "データベースへの学生情報削除時に例外が発生しました。具体的な原因：" + str(e)})


def delete_students(request):
    """複数の学生情報を削除する"""
    # フロントエンドから送られてきた値を受け取る
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 送られてきた集合を順に処理する
        for one_student in data['student']:
            # 現在の記録を検索する
            obj_student = Student.objects.get(sno=one_student['sno'])
            # 削除を実行する
            obj_student.delete()
        # ORMを使用してすべての学生情報を取得し、オブジェクトを辞書形式に変換する
        obj_students = Student.objects.all().values()
        # 外側のコンテナをListに変換する
        students = list(obj_students)
        # 戻り値
        return JsonResponse({'code': 1, 'data': students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "データベースへの複数学生情報削除時に例外が発生しました。具体的な原因：" + str(e)})


def upload(request):
    """アップロードされたファイルを受け取る"""
    # アップロードされたファイルを受け取る
    rev_file = request.FILES.get('avatar')
    # ファイルが存在するかどうかを確認する
    if not rev_file:
        return JsonResponse({'code':0, 'msg':'画像が存在しません！'})
    # ユニークな名前を取得する： uuid +hash
    new_name = get_random_str()
    # 書き込み用のURLを準備する
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1] )
    # ディスクに書き込みを開始する
    try:
        f = open(file_path,'wb')
        # 複数回にわたって書き込む
        for i in rev_file.chunks():
            f.write(i)
        # 閉じる必要がある
        f.close()
        # 戻り値
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(rev_file.name)[1]})

    except Exception as e:
        return JsonResponse({'code':0, 'msg':str(e)})


def get_random_str():
    """uuidのランダム数を取得する"""
    uuid_val = uuid.uuid4()
    """uuidのランダム数文字列を取得する"""
    uuid_str = str(uuid_val).encode('utf-8')
    """md5インスタンスを取得する"""
    md5 = hashlib.md5()
    """uuidのmd5ダイジェストを取得する"""
    md5.update(uuid_str)
    """固定長の文字列を返す"""
    return md5.hexdigest()



def import_students_excel(request):
    """Excelから学生情報を一括でインポートする"""
    # ========1. ExcelファイルをMediaフォルダに受け取り保存 =======
    rev_file = request.FILES.get('excel')
    # ファイルの存在を判断
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': 'Excelファイルが存在しません！'})
    # ユニークな名前を生成：uuid + ハッシュ
    new_name = get_random_str()
    # 書き込み用URLを準備
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    # ディスクへの書き込みを開始
    try:
        f = open(file_path, 'wb')
        # 複数回にわたって書き込む
        for i in rev_file.chunks():
            f.write(i)
        # 閉じる必要がある
        f.close()
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})

    #====== 2. Mediaフォルダに保存されたデータを読み取る =====
    ex_students = read_excel_dict(file_path)

    # ====3. 読み取ったデータをデータベースに保存する ====
    # 変数を定義：成功： success, 失敗： error, エラーのsno： error_snos
    success = 0
    error = 0
    error_snos = []

    # 反復処理を開始
    for one_student in ex_students:
        try:
            obj_student = Student.objects.create(sno=one_student['sno'], name=one_student['name'], gender=one_student['gender'],
                                                 birthday=one_student['birthday'], mobile=one_student['mobile'],
                                                 email=one_student['email'], address=one_student['address'])
            # カウント
            success += 1
        except:
            # 失敗した場合
            error += 1
            error_snos.append(one_student['sno'])

    # 4. インポート情報（成功：success, 失敗：error--（sno））, 全学生情報を返す
    obj_students = Student.objects.all().values()
    students = list(obj_students)
    return JsonResponse({'code':1, 'success':success, 'error':error, 'errors':error_snos, 'data':students})

def export_student_excel(request):
    """データをExcelにエクスポートする"""
    # すべての学生情報を取得
    obj_students = Student.objects.all().values()
    # Listに変換
    students = list(obj_students)
    # ファイル名を準備
    excel_name = get_random_str() + ".xlsx"
    # 書き込みパスを準備
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # Excelに書き込む
    write_to_excel(students, path)
    # 返す
    return JsonResponse({'code':1, 'name':excel_name})

def read_excel_dict(path: str):
    """Excelデータを読み取り、辞書に保存 --- [{},{},{},]"""
    # workbookのインスタンス化
    workbook = openpyxl.load_workbook(path)
    # sheetのインスタンス化
    sheet = workbook['student']
    # 最終データを保存する変数の定義--[]
    students = []
    # keyを準備
    keys = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address']
    # 反復処理
    for row in sheet.iter_rows(min_row=2):  # 最初の行がヘッダーであると仮定
        # 一時的な辞書を定義
        temp_dict = {}
        # 値とkeyを組み合わせる
        for index, cell in enumerate(row):
            if index < len(keys):  # indexがkeyの範囲内であることを確認
                # 組み合わせる
                temp_dict[keys[index]] = cell.value
        # listに追加する
        students.append(temp_dict)
    # 返す
    return students

def write_to_excel(data:list, path:str):
    """データベースの内容をExcelに書き込む"""
    # workbookをインスタンス化
    workbook = openpyxl.Workbook()
    # sheetをアクティブにする
    sheet = workbook.active
    # sheetに名前をつける
    sheet.title = 'student'
    # keysを準備
    keys = data[0].keys()
    # データを書き込む準備
    for index, item in enumerate(data):
        # 各要素を反復処理する
        for k,v in enumerate(keys):
            sheet.cell(row=index + 1, column=k+ 1, value=str(item[v]))
    # ファイルに書き込む
    workbook.save(path)